"""
Portfolio Export Module
Exports portfolio data to PDF and Excel formats
"""

from datetime import datetime
from io import BytesIO
from typing import List, Dict, Any, Optional
import pandas as pd

# PDF exports using reportlab
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        PageBreak, Image, KeepTogether
    )
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

# Excel exports using openpyxl via pandas
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class PortfolioExporter:
    """Export portfolio data to various formats"""

    def __init__(self):
        self.styles = getSampleStyleSheet() if REPORTLAB_AVAILABLE else None

    def export_to_excel(
        self,
        portfolio: Dict[str, Any],
        projects: List[Dict[str, Any]],
        metrics: Optional[Dict[str, Any]] = None,
        risks: Optional[List[Dict[str, Any]]] = None
    ) -> BytesIO:
        """
        Export portfolio to Excel with multiple sheets

        Args:
            portfolio: Portfolio data
            projects: List of projects in portfolio
            metrics: Portfolio metrics (optional)
            risks: Portfolio risks (optional)

        Returns:
            BytesIO object containing Excel file
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

        output = BytesIO()

        # Create Excel writer
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Sheet 1: Portfolio Summary
            summary_data = {
                'Field': [
                    'Portfolio Name',
                    'Description',
                    'Total Budget',
                    'Total Capacity (FTE)',
                    'Created At',
                    'Total Projects',
                    'Budget Allocated',
                    'Capacity Allocated',
                ],
                'Value': [
                    portfolio.get('name', 'N/A'),
                    portfolio.get('description', 'N/A'),
                    f"R$ {portfolio.get('total_budget', 0):,.2f}",
                    f"{portfolio.get('total_capacity', 0):.1f}",
                    portfolio.get('created_at', 'N/A'),
                    len(projects),
                    f"R$ {sum(p.get('budget_allocated', 0) for p in projects):,.2f}",
                    f"{sum(p.get('capacity_allocated', 0) for p in projects):.1f}",
                ]
            }
            df_summary = pd.DataFrame(summary_data)
            df_summary.to_excel(writer, sheet_name='Portfolio Summary', index=False)

            # Sheet 2: Projects
            if projects:
                projects_data = []
                for p in projects:
                    projects_data.append({
                        'Project ID': p.get('project_id', 'N/A'),
                        'Project Name': p.get('project_name', 'N/A'),
                        'Priority': p.get('priority', 'N/A'),
                        'Business Value': p.get('business_value', 0),
                        'WSJF Score': p.get('wsjf_score', 0),
                        'CoD Weekly (R$)': p.get('cod_weekly', 0),
                        'Budget Allocated (R$)': p.get('budget_allocated', 0),
                        'Capacity Allocated (FTE)': p.get('capacity_allocated', 0),
                        'Time Criticality': p.get('time_criticality', 0),
                        'Risk Reduction': p.get('risk_reduction', 0),
                        'Status': p.get('status', 'N/A'),
                    })
                df_projects = pd.DataFrame(projects_data)
                df_projects.to_excel(writer, sheet_name='Projects', index=False)

            # Sheet 3: Metrics (if provided)
            if metrics:
                metrics_data = {
                    'Metric': [],
                    'Value': []
                }
                for key, value in metrics.items():
                    # Format metric name
                    metric_name = key.replace('_', ' ').title()
                    metrics_data['Metric'].append(metric_name)

                    # Format value
                    if isinstance(value, float):
                        if 'budget' in key.lower() or 'cost' in key.lower():
                            metrics_data['Value'].append(f"R$ {value:,.2f}")
                        elif 'capacity' in key.lower() or 'fte' in key.lower():
                            metrics_data['Value'].append(f"{value:.1f}")
                        else:
                            metrics_data['Value'].append(f"{value:.2f}")
                    else:
                        metrics_data['Value'].append(str(value))

                df_metrics = pd.DataFrame(metrics_data)
                df_metrics.to_excel(writer, sheet_name='Metrics', index=False)

            # Sheet 4: Risks (if provided)
            if risks:
                risks_data = []
                for r in risks:
                    risks_data.append({
                        'Risk ID': r.get('id', 'N/A'),
                        'Title': r.get('risk_title', 'N/A'),
                        'Category': r.get('risk_category', 'N/A'),
                        'Probability (1-5)': r.get('probability', 0),
                        'Impact (1-5)': r.get('impact', 0),
                        'Risk Score': r.get('risk_score', 0),
                        'Status': r.get('status', 'N/A'),
                        'Owner': r.get('owner', 'N/A'),
                        'Mitigation Plan': r.get('mitigation_plan', 'N/A'),
                    })
                df_risks = pd.DataFrame(risks_data)
                df_risks.to_excel(writer, sheet_name='Risks', index=False)

            # Apply formatting to all sheets
            self._format_excel_sheets(writer.book)

        output.seek(0)
        return output

    def _format_excel_sheets(self, workbook):
        """Apply formatting to Excel workbook"""
        # Define styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]

            # Format headers
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Apply borders to all cells with data
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.border = border

    def export_to_pdf(
        self,
        portfolio: Dict[str, Any],
        projects: List[Dict[str, Any]],
        metrics: Optional[Dict[str, Any]] = None,
        risks: Optional[List[Dict[str, Any]]] = None,
        include_charts: bool = False
    ) -> BytesIO:
        """
        Export portfolio to PDF report

        Args:
            portfolio: Portfolio data
            projects: List of projects in portfolio
            metrics: Portfolio metrics (optional)
            risks: Portfolio risks (optional)
            include_charts: Whether to include charts (optional)

        Returns:
            BytesIO object containing PDF file
        """
        if not REPORTLAB_AVAILABLE:
            raise ImportError("reportlab is required for PDF export. Install with: pip install reportlab")

        output = BytesIO()

        # Create PDF document
        doc = SimpleDocTemplate(
            output,
            pagesize=letter,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=18
        )

        # Container for PDF elements
        elements = []

        # Define styles
        styles = self.styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#366092'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#366092'),
            spaceAfter=12,
            spaceBefore=12
        )

        # Title
        title = Paragraph(f"Portfolio Report: {portfolio.get('name', 'N/A')}", title_style)
        elements.append(title)

        # Date
        date_str = datetime.now().strftime("%Y-%m-%d %H:%M")
        date_para = Paragraph(f"<i>Generated on: {date_str}</i>", styles['Normal'])
        elements.append(date_para)
        elements.append(Spacer(1, 20))

        # Portfolio Summary
        elements.append(Paragraph("Portfolio Summary", heading_style))

        summary_data = [
            ['Field', 'Value'],
            ['Portfolio Name', portfolio.get('name', 'N/A')],
            ['Description', portfolio.get('description', 'N/A')[:100]],
            ['Total Budget', f"R$ {portfolio.get('total_budget', 0):,.2f}"],
            ['Total Capacity', f"{portfolio.get('total_capacity', 0):.1f} FTE"],
            ['Total Projects', str(len(projects))],
            ['Budget Allocated', f"R$ {sum(p.get('budget_allocated', 0) for p in projects):,.2f}"],
            ['Capacity Allocated', f"{sum(p.get('capacity_allocated', 0) for p in projects):.1f} FTE"],
        ]

        summary_table = Table(summary_data, colWidths=[2.5*inch, 4*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 20))

        # Projects
        if projects:
            elements.append(PageBreak())
            elements.append(Paragraph("Projects", heading_style))

            projects_data = [['Project', 'Priority', 'Business Value', 'WSJF', 'Budget (R$)', 'Capacity (FTE)']]
            for p in projects:
                projects_data.append([
                    p.get('project_name', 'N/A')[:30],
                    str(p.get('priority', 'N/A')),
                    str(p.get('business_value', 0)),
                    f"{p.get('wsjf_score', 0):.1f}",
                    f"{p.get('budget_allocated', 0):,.0f}",
                    f"{p.get('capacity_allocated', 0):.1f}",
                ])

            projects_table = Table(projects_data, colWidths=[2*inch, 0.8*inch, 1*inch, 0.8*inch, 1.2*inch, 1*inch])
            projects_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            elements.append(projects_table)
            elements.append(Spacer(1, 20))

        # Metrics
        if metrics:
            elements.append(PageBreak())
            elements.append(Paragraph("Portfolio Metrics", heading_style))

            metrics_data = [['Metric', 'Value']]
            for key, value in metrics.items():
                metric_name = key.replace('_', ' ').title()

                # Format value
                if isinstance(value, float):
                    if 'budget' in key.lower() or 'cost' in key.lower():
                        formatted_value = f"R$ {value:,.2f}"
                    elif 'capacity' in key.lower() or 'fte' in key.lower():
                        formatted_value = f"{value:.1f}"
                    else:
                        formatted_value = f"{value:.2f}"
                else:
                    formatted_value = str(value)

                metrics_data.append([metric_name, formatted_value])

            metrics_table = Table(metrics_data, colWidths=[3*inch, 3.5*inch])
            metrics_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            elements.append(metrics_table)
            elements.append(Spacer(1, 20))

        # Risks
        if risks:
            elements.append(PageBreak())
            elements.append(Paragraph("Portfolio Risks", heading_style))

            risks_data = [['Risk Title', 'Category', 'Prob.', 'Impact', 'Score', 'Status']]
            for r in risks:
                risks_data.append([
                    r.get('risk_title', 'N/A')[:40],
                    r.get('risk_category', 'N/A'),
                    str(r.get('probability', 0)),
                    str(r.get('impact', 0)),
                    str(r.get('risk_score', 0)),
                    r.get('status', 'N/A'),
                ])

            risks_table = Table(risks_data, colWidths=[2.5*inch, 1*inch, 0.6*inch, 0.6*inch, 0.6*inch, 1.2*inch])
            risks_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            elements.append(risks_table)

        # Build PDF
        doc.build(elements)

        output.seek(0)
        return output


def export_portfolio_excel(portfolio, projects, metrics=None, risks=None):
    """Convenience function to export portfolio to Excel"""
    exporter = PortfolioExporter()
    return exporter.export_to_excel(portfolio, projects, metrics, risks)


def export_portfolio_pdf(portfolio, projects, metrics=None, risks=None):
    """Convenience function to export portfolio to PDF"""
    exporter = PortfolioExporter()
    return exporter.export_to_pdf(portfolio, projects, metrics, risks)
